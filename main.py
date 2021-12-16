# -*- coding: utf-8 -*-
#
# Generate tags on product from excel
#
# Edit by Kochetkov Artem
# skype: artemk_85
# mail: kochetkov1985@mail.ru
#

from __future__ import absolute_import, division, print_function, unicode_literals
from io import BytesIO
import datetime
import os

import upcean
import openpyxl
import xlsxwriter
from barcode import EAN13
from barcode.writer import ImageWriter
import constant


start_row = constant.START_ROW  # В экселе +1
start_col = constant.START_COL  # 'C'
end_col = constant.END_COL   # 'T'
col_range = constant.COL_RANGE
start_row_read = constant.START_ROW_READ
end_row_read = constant.END_ROW_READ
# col A - BE или 1 - 58


def get_data_from_xls(fn: str):
    data = []
    xls_file = os.path.abspath(fn)
    wb = openpyxl.load_workbook(filename=xls_file, read_only=True)
    ws = wb['Лист1']

    for row in range(start_row_read, end_row_read):
        if ws.cell(row, 2).value == None and ws.cell(row, 3).value == None and ws.cell(row, 4).value == None and ws.cell(row, 5).value == None and ws.cell(row, 6).value == None and ws.cell(row, 7).value == None:
            print(f'End data in Excel in {row} row.')
            break

        data_row = []
        for col in range(1, 58):
            data_row.append(ws.cell(row=row, column=col).value)
        data.append(data_row)
    return data

if __name__ == "__main__":
    all_file = 'all_data.xlsx'
    barcode_file = f'barcode_all_{datetime.datetime.date(datetime.datetime.now())}.xlsx'

    xsl_data = get_data_from_xls(all_file)
    # xsl_data = [[43, 180, 'Кольцо обручальное, БК, 18 (ш5)', 'СПЕЦ', 'DЕ', 'ЗОЛОТО', 585, '18,0', '125000', 4300128141, 'КРАСНЫЙ', 'БЕЗ ПОКРЫТИЯ', '2200001231238', 20320, 8000, 20320, 15749, None, '18,0', 585, None, 4300128141, '2200001231238', '05.06.2021', None, None, None, None, '13.04.2021', '4300013266', None, 48, 2.54, 2.54, 121.92, 121.92, 0, 'RUB', 0, 54.17, 'Г', 1, 0, 4318.29, 62.7, 0, 0, 6604.41, 1.5, 2.54, 2.54, 121.92, 121.92, 72.398, None, 'ОБРУЧ', 'БЕЗ АЛМ.ОГРАНКИ']]
    # print(xsl_data)

    doc = xlsxwriter.Workbook(filename=barcode_file)

    for elem in xsl_data:
        doc_ws = doc.add_worksheet(str(elem[12]))
        doc_ws.set_zoom(235)
        doc_ws.set_column('A:A', 0.83)
        doc_ws.set_column(col_range, 0.415)
        doc_ws.set_default_row(5.25)

        doc_ws.set_row_pixels(start_row - 2, 4)
        doc_ws.set_row_pixels(start_row - 1, 4)
        doc_ws.set_row_pixels(start_row, 7)
        doc_ws.set_row_pixels(start_row + 1, 26)
        doc_ws.set_row_pixels(start_row + 2, 9)
        doc_ws.set_row_pixels(start_row + 3, 11)
        doc_ws.set_row_pixels(start_row + 10, 11)
        doc_ws.set_row_pixels(start_row + 11, 13)
        doc_ws.set_row_pixels(start_row + 12, 9)
        doc_ws.set_row_pixels(start_row + 13, 10)
        doc_ws.set_row_pixels(start_row + 14, 15)
        doc_ws.set_row_pixels(start_row + 15, 15)
        doc_ws.set_row_pixels(start_row + 16, 5)
        doc_ws.set_row_pixels(start_row + 17, 15)
        doc_ws.set_row_pixels(start_row + 18, 12)

        doc_ws.print_area('B1:S24')

        # Название ООО "ОРО"
        type_met_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 5,
                'text_h_align': 2,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row, start_col, start_row, start_col + 17, 'ООО "ОРО"', type_met_format)  # 'B6:H6'

        # Наименование изделия
        name_format = doc.add_format(
            {
                'text_wrap': True,
                'font_name': 'Times New Roman',
                'font_size': 7,
                'text_h_align': 2,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+1, start_col, start_row+1, start_col+17, elem[2], name_format)

        # Название товарного направления
        ntn_format = doc.add_format({
                'font_name': 'Arial',
                'font_size': 7,
                'text_h_align': 1,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+2, start_col, start_row+2, start_col+11, elem[3], ntn_format)

        # Коллекция
        koll_format = doc.add_format({
                'font_name': 'Arial',
                'font_size': 7,
                'text_h_align': 3,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+2, start_col+14, start_row+2, start_col+17, elem[4], koll_format)  # 'P13:S13'

        # Тип металла
        type_met_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 5,
                'text_h_align': 2,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row + 3, start_col, start_row + 3, start_col+6, elem[5], type_met_format)  # 'B14:H14'

        # Проба
        proba_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 8,
                'text_h_align': 2,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+3, start_col+7, start_row+3, start_col+10, elem[6], proba_format)  # 'I14:L14'

        # Размер
        raz_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 5,
                'text_h_align': 2,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+3, start_col+11, start_row+3, start_col+13, 'Р-р', raz_format)

        raz1_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 8,
                'text_h_align': 2,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+3, start_col+14, start_row+3, start_col+17, elem[7], raz1_format)

        # Артикул
        art_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 5,
                'text_h_align': 1,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+4, start_col, start_row+4, start_col+17, f'Арт {elem[8]}', art_format)

        # САП код
        sap_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 5,
                'text_h_align': 1,  # left
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+5, start_col, start_row+5, start_col+17, f'САП код: {elem[9]}', sap_format)

        # Цвет металла
        cvet_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 5,
                'text_h_align': 3,  # right
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+6, start_col, start_row+6, start_col+17, f'Цв. {elem[10]}', cvet_format)

        # Покрытие


        # Штрихкод
        image_data = BytesIO()
        ean = EAN13(str(elem[12]), writer=ImageWriter())
        ean.write(image_data, options={"write_text": False})

        x_scale = 0.21  # *13.89
        y_scale = 0.115  # *5.39

        doc_ws.insert_image(
            f'A{start_row+8}',
            str(elem[12]),
            {
                'image_data': image_data,
                'x_offset': 0,
                'y_offset': 0,
                'x_scale': x_scale,
                'y_scale': y_scale,
                'object_position': 0,
            }
        )

        bc = doc.add_format({
                'font_name': 'Arial',
                'font_size': 7,
                'text_h_align': 2,
                'text_v_align': 1,  # top
            }
        )
        doc_ws.merge_range(start_row+10, start_col, start_row+10, start_col+17, str(elem[12]), bc)

        # Цена
        cena_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 8,
                'text_h_align': 2,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+11, start_col, start_row+11, start_col+4, 'Цена:', cena_format)

        cena2_format = doc.add_format({
                'bold': True,
                'font_name': 'Times New Roman',
                'font_size': 10,
                'text_h_align': 2,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+11, start_col+5, start_row+11, start_col+13, elem[13], cena2_format)

        cena3_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 8,
                'text_h_align': 2,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+11, start_col+14, start_row+11, start_col+17, 'руб.', cena3_format)

        # Цена за грамм
        cena_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 5,
                'text_h_align': 2,
                'text_v_align': 1,
            }
        )
        doc_ws.merge_range(start_row+12, start_col, start_row+12, start_col+4, 'За гр.:', cena_format)

        cena2_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 5,
                'text_h_align': 2,
                'text_v_align': 1,
            }
        )
        doc_ws.merge_range(start_row+12, start_col+5, start_row+12, start_col+13, elem[14], cena2_format)

        cena3_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 5,
                'text_h_align': 2,
                'text_v_align': 1,
            }
        )
        doc_ws.merge_range(start_row+12, start_col+14, start_row+12, start_col+17, 'руб.', cena3_format)

        # Размер
        raz1_format = doc.add_format({
                'font_name': 'Arial',
                'font_size': 7,
                'text_h_align': 2,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+14, start_col, start_row+14, start_col+4, elem[18], raz1_format)

        # Цена прочерк
        cena_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 9,
                'text_h_align': 2,
                'text_v_align': 2,
                'font_strikeout': 1,
            }
        )
        doc_ws.merge_range(start_row+14, start_col+10, start_row+14, start_col+17, f'{elem[15]}p', cena_format)

        # Проба
        if elem[19] == 375:
            proba_format = doc.add_format({
                    'font_name': 'Arial',
                    'font_size': 7,
                    'text_h_align': 2,
                    'text_v_align': 2,
                }
            )
            doc_ws.merge_range(start_row+15, start_col, start_row+15, start_col+4, elem[19], proba_format)  # 'B25:F25'

        # Цена продаж
        cena_format = doc.add_format({
                'font_name': 'Times New Roman',
                'font_size': 9,
                'text_h_align': 2,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+15, start_col+10, start_row+15, start_col+17, f'{elem[16]}p', cena_format)

        # Штрихкод
        image_data2 = BytesIO()
        ean2 = EAN13(str(elem[22]), writer=ImageWriter())
        ean2.write(image_data2, options={"write_text": False, "includetext": True})

        x_scale = 0.21  # *13.89
        y_scale = 0.076  # *5.39

        doc_ws.insert_image(
            start_row+17,
            0,
            f'{elem[22]}.png',
            {
                'image_data': image_data2,
                'x_offset': 0,
                'y_offset': 0,
                'x_scale': x_scale,
                'y_scale': y_scale,
                'object_position': 0,
            }
        )

        # САП код
        sap_format = doc.add_format({
                'font_name': 'Arial',
                'font_size': 8,
                'text_h_align': 2,
                'text_v_align': 2,
            }
        )
        doc_ws.merge_range(start_row+18, start_col, start_row+18, start_col+17, elem[9], sap_format)

    doc.close()
